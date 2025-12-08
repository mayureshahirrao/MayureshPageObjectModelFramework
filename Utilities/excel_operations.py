import openpyxl

def read_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []
    for row in sheet.iter_rows():
        data.append(row)

    return data

def create_excel(file_path, sheet_name, data):
    workbook = openpyxl.Workbook(file_path)
    sheet = workbook.active
    sheet.title = sheet_name

    for row in data:
        sheet.append(row)

    workbook.save(file_path)

def write_excel(file_path, sheet_name, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for row in data:
        sheet.append(row)

    workbook.save(file_path)

def get_row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    return sheet.max_row

def get_column_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    return sheet.max_column

def get_cell_value(file_path, sheet_name, row, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    return sheet.cell(row=row, column=column).value.strip()

def set_cell_value(file_path, sheet_name, row, column, value):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    sheet.cell(row=row, column=column).value = value
    workbook.save(file_path)

if __name__ =="__main__":
    read_file_path = '..\\Excel\\testdata.xlsx'
    sheet_name = 'LoginTest'
    row_count = get_row_count(read_file_path, sheet_name)
    column_count = get_column_count(read_file_path, sheet_name)
    user_list = []

    for row in range(2, row_count + 1):
        user = []
        for column in range(1, column_count + 1):
            user.append(get_cell_value(read_file_path, sheet_name, row, column))
        user_list.append(user)

    print(user_list)